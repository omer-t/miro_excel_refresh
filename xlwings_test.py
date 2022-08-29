import pandas as pd
import xlwings as xw

import modules.snowflake as sf


# conn = sf.snowflake_connection('credentials/snowflake.json')
# df = sf.sql_file_to_dataframe('sql/example.sql', conn)
# sql = 'select * from production.marketing.paid_search limit 100'
# df = pd.read_sql(sql, conn)

# wb = xw.Book()
# wb = xw.Book('no-file.xlsx')  # this will create a new workbook
app_excel = xw.App(visible = False)
wb = xw.Book('excel/paid_search.xlsx')
wb.api.RefreshAll()

wb.api.


ws = wb.sheets('1d')

pivot_cache = wb.api.ActiveSheet.PivotTables('PivotTable1').PivotCache()
pivot_cache = ws.api.PivotTables('PivotTable1').PivotCache()

ws.api.PivotTables('1d').PivotCache()

for idx, pt in enumerate(ws.api.PivotTables()):
    print(pt.Name)


wb.sheets['7d'].select()
wb.api.ActiveSheet.PivotTables("PivotTable1").PivotCache().refresh()

wb.sheets('7d').PivotTables("PivotTable1").RefreshTable

# ---- openpyxl

from openpyxl import Workbook, load_workbook

workbook = load_workbook('excel/report.xlsx')
print(workbook)
worksheet = workbook['7d']

pivot_table = [p for p in worksheet._pivots if p.name == 'PivotTable1'][0]

# ----
wb.RefreshAll()


# -----
import openpyxl
wb_obj = openpyxl.load_workbook('excel/paid_search.xlsx')
sheet_obj = wb_obj.active
sheet_obj.PivotTables("1d").PivotCache.Refresh
Worksheets("SheetName").PivotTables("PivotTableName").PivotCache.Refresh

